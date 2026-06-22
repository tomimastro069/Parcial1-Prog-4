import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import select
from sqlalchemy.orm import selectinload, joinedload

from app.Core.UnitOfWork.unit_of_work import UnitOfWork
from app.Modules.Producto.Model.producto import Producto
from app.Modules.Producto.Model.productoIngrediente import ProductoIngrediente
from app.Modules.Categoria.categoria import Categoria
from app.Modules.Ingrediente.ingrediente import Ingrediente
from app.Modules.Pedidos.Model.pedido import Pedido

class ReporteService:
    def __init__(self, uow: UnitOfWork):
        self.uow = uow

    def _style_sheet(self, ws, headers: list[str]):
        """Agrega los headers estilizados a la hoja."""
        ws.append(headers)
        
        # Color azul oscuro institucional (#1F3864) con texto blanco y negrita
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        ws.row_dimensions[1].height = 25

    def _auto_adjust_columns(self, ws):
        """Ajusta automáticamente el ancho de las columnas basado en el contenido."""
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    def _save_workbook(self, wb: Workbook) -> io.BytesIO:
        """Guarda el workbook en un buffer de memoria y lo retorna."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generar_reporte_general(self) -> io.BytesIO:
        """Genera un archivo Excel multi-hoja con toda la información del sistema."""
        wb = Workbook()
        
        # Hoja Pedidos
        ws_pedidos = wb.active
        ws_pedidos.title = "Pedidos"
        self._style_sheet(
            ws_pedidos, 
            ["ID", "Fecha", "Estado", "Forma de pago", "Subtotal", "Costo envío", "Total", "Notas"]
        )
        with self.uow:
            stmt = select(Pedido).options(
                selectinload(Pedido.estado),
                selectinload(Pedido.forma_pago)
            ).order_by(Pedido.created_at.desc())
            pedidos = self.uow.session.exec(stmt).all()
            for p in pedidos:
                fecha_str = p.created_at.strftime("%d/%m/%Y %H:%M") if p.created_at else ""
                estado_desc = p.estado.descripcion if p.estado else p.estado_codigo
                forma_pago_desc = p.forma_pago.descripcion if p.forma_pago else p.forma_pago_codigo
                ws_pedidos.append([
                    p.id,
                    fecha_str,
                    estado_desc,
                    forma_pago_desc,
                    float(p.subtotal),
                    float(p.costo_envio),
                    float(p.total),
                    p.notas or ""
                ])
        self._auto_adjust_columns(ws_pedidos)

        # Hoja Productos
        ws_productos = wb.create_sheet(title="Productos")
        self._style_sheet(
            ws_productos,
            ["Nombre", "Precio", "Descripción", "Categorías", "Ingredientes", "Terminado", "Estado"]
        )
        with self.uow:
            stmt = select(Producto).options(
                selectinload(Producto.categoria),
                selectinload(Producto.producto_ingredientes).joinedload(ProductoIngrediente.ingrediente)
            ).order_by(Producto.nombre)
            productos = self.uow.session.exec(stmt).all()
            for p in productos:
                categoria_nombre = self.uow.categorias.get_full_path(p.categoria_id) if p.categoria_id else ""
                ingredientes_nombres = [
                    pi.ingrediente.nombre for pi in p.producto_ingredientes if pi.ingrediente
                ]
                
                ws_productos.append([
                    p.nombre,
                    float(p.precio_base),
                    p.descripcion or "",
                    categoria_nombre,
                    ", ".join(ingredientes_nombres),
                    "Sí" if p.es_terminado else "No",
                    "Activo" if p.is_active else "Inactivo"
                ])
        self._auto_adjust_columns(ws_productos)

        # Hoja Categorías
        ws_categorias = wb.create_sheet(title="Categorías")
        self._style_sheet(
            ws_categorias,
            ["Nombre", "Descripción", "Categoría padre", "Estado"]
        )
        with self.uow:
            stmt = select(Categoria).options(
                selectinload(Categoria.parent)
            ).order_by(Categoria.nombre)
            categorias = self.uow.session.exec(stmt).all()
            for c in categorias:
                nombre_corto = c.nombre.split(" / ")[-1] if c.nombre else ""
                parent_desc = ""
                if c.parent:
                    parent_desc = c.parent.nombre.split(" / ")[-1]
                
                ws_categorias.append([
                    nombre_corto,
                    c.descripcion or "",
                    parent_desc,
                    "Activa" if c.is_active else "Inactiva"
                ])
        self._auto_adjust_columns(ws_categorias)

        # Hoja Ingredientes
        ws_ingredientes = wb.create_sheet(title="Ingredientes")
        self._style_sheet(
            ws_ingredientes,
            ["Nombre", "Descripción", "Alérgeno", "Estado"]
        )
        with self.uow:
            ingredientes = self.uow.session.exec(select(Ingrediente).order_by(Ingrediente.nombre)).all()
            for ing in ingredientes:
                ws_ingredientes.append([
                    ing.nombre,
                    ing.descripcion or "",
                    "Sí" if ing.es_alergeno else "No",
                    "Activo" if ing.is_active else "Inactivo"
                ])
        self._auto_adjust_columns(ws_ingredientes)

        return self._save_workbook(wb)

    def generar_reporte_productos(self) -> io.BytesIO:
        """Genera un archivo Excel con una sola hoja conteniendo la lista de Productos."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        self._style_sheet(
            ws,
            ["Nombre", "Precio", "Descripción", "Categorías", "Ingredientes", "Terminado", "Estado"]
        )
        with self.uow:
            stmt = select(Producto).options(
                selectinload(Producto.categoria),
                selectinload(Producto.producto_ingredientes).joinedload(ProductoIngrediente.ingrediente)
            ).order_by(Producto.nombre)
            productos = self.uow.session.exec(stmt).all()
            for p in productos:
                categoria_nombre = self.uow.categorias.get_full_path(p.categoria_id) if p.categoria_id else ""
                ingredientes_nombres = [
                    pi.ingrediente.nombre for pi in p.producto_ingredientes if pi.ingrediente
                ]
                
                ws.append([
                    p.nombre,
                    float(p.precio_base),
                    p.descripcion or "",
                    categoria_nombre,
                    ", ".join(ingredientes_nombres),
                    "Sí" if p.es_terminado else "No",
                    "Activo" if p.is_active else "Inactivo"
                ])
        self._auto_adjust_columns(ws)
        return self._save_workbook(wb)

    def generar_reporte_categorias(self) -> io.BytesIO:
        """Genera un archivo Excel con una sola hoja conteniendo la lista de Categorías."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Categorías"
        self._style_sheet(
            ws,
            ["Nombre", "Descripción", "Categoría padre", "Estado"]
        )
        with self.uow:
            stmt = select(Categoria).options(
                selectinload(Categoria.parent)
            ).order_by(Categoria.nombre)
            categorias = self.uow.session.exec(stmt).all()
            for c in categorias:
                nombre_corto = c.nombre.split(" / ")[-1] if c.nombre else ""
                parent_desc = ""
                if c.parent:
                    parent_desc = c.parent.nombre.split(" / ")[-1]
                
                ws.append([
                    nombre_corto,
                    c.descripcion or "",
                    parent_desc,
                    "Activa" if c.is_active else "Inactiva"
                ])
        self._auto_adjust_columns(ws)
        return self._save_workbook(wb)

    def generar_reporte_ingredientes(self) -> io.BytesIO:
        """Genera un archivo Excel con una sola hoja conteniendo la lista de Ingredientes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingredientes"
        self._style_sheet(
            ws,
            ["Nombre", "Descripción", "Alérgeno", "Estado"]
        )
        with self.uow:
            ingredientes = self.uow.session.exec(select(Ingrediente).order_by(Ingrediente.nombre)).all()
            for ing in ingredientes:
                ws.append([
                    ing.nombre,
                    ing.descripcion or "",
                    "Sí" if ing.es_alergeno else "No",
                    "Activo" if ing.is_active else "Inactivo"
                ])
        self._auto_adjust_columns(ws)
        return self._save_workbook(wb)

    def generar_reporte_pedidos(self) -> io.BytesIO:
        """Genera un archivo Excel con una sola hoja conteniendo la lista de Pedidos."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        self._style_sheet(
            ws,
            ["ID", "Fecha", "Estado", "Forma de pago", "Subtotal", "Costo envío", "Total", "Notas"]
        )
        with self.uow:
            stmt = select(Pedido).options(
                selectinload(Pedido.estado),
                selectinload(Pedido.forma_pago)
            ).order_by(Pedido.created_at.desc())
            pedidos = self.uow.session.exec(stmt).all()
            for p in pedidos:
                fecha_str = p.created_at.strftime("%d/%m/%Y %H:%M") if p.created_at else ""
                estado_desc = p.estado.descripcion if p.estado else p.estado_codigo
                forma_pago_desc = p.forma_pago.descripcion if p.forma_pago else p.forma_pago_codigo
                ws.append([
                    p.id,
                    fecha_str,
                    estado_desc,
                    forma_pago_desc,
                    float(p.subtotal),
                    float(p.costo_envio),
                    float(p.total),
                    p.notes or "" if hasattr(p, 'notes') else p.notas or ""
                ])
        self._auto_adjust_columns(ws)
        return self._save_workbook(wb)
